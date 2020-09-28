from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementNotInteractableException
from openpyxl import load_workbook
from openpyxl import Workbook


url = 'https://www.tvtc.gov.sa/Arabic/Departments/Departments/pt/Query/Pages/OQ.aspx'
driver = webdriver.Chrome(executable_path="E:/test/chromedriver.exe")
driver.maximize_window()
driver.get(url)
iframeElement = driver.find_element_by_tag_name('iframe')
driver.switch_to.frame(iframeElement)
women = driver.find_element_by_xpath('//*[@id="table2"]/tbody/tr/td[1]/font/span/a')
women.click()
cities = [city.get_attribute('href') for city in driver.find_elements_by_xpath('//*[@id="table2"]/tbody/tr/td/font/a')]
sh=1
wb = load_workbook('test.xlsx')
for city in cities:
    sheet = wb.create_sheet('Sheet{0}'.format(sh))
    row = 1
    print(city)
    driver.get(city)
    inner_cities = [c.get_attribute('href') for c in driver.find_elements_by_xpath('//*[@id="table2"]/tbody/tr/td/font/span/a')]
    for inner_city in inner_cities:
        print(inner_city)
        driver.get(inner_city)
        tr = driver.find_elements_by_xpath('/html/body/div/table/tbody//tr')
        i = 2
        for t in tr[2:]:
            name = t.find_element_by_xpath('/html/body/div/table/tbody/tr[{0}]/td[9]'.format(i)).text
            kind = t.find_element_by_xpath('/html/body/div/table/tbody/tr[{0}]/td[8]'.format(i)).text
            tele = t.find_element_by_xpath('/html/body/div/table/tbody/tr[{0}]/td[7]'.format(i)).text
            fax = t.find_element_by_xpath('/html/body/div/table/tbody/tr[{0}]/td[6]'.format(i)).text
            district = t.find_element_by_xpath('/html/body/div/table/tbody/tr[{0}]/td[5]'.format(i)).text
            address = t.find_element_by_xpath('/html/body/div/table/tbody/tr[{0}]/td[4]'.format(i)).text
            licence = t.find_element_by_xpath('/html/body/div/table/tbody/tr[{0}]/td[3]'.format(i)).text
            supervisor = t.find_element_by_xpath('/html/body/div/table/tbody/tr[{0}]/td[2]'.format(i)).text
            sheet.cell(row=row, column=1, value=name)
            sheet.cell(row=row, column=2, value=kind)
            sheet.cell(row=row, column=3, value=tele)
            sheet.cell(row=row, column=4, value=fax)
            sheet.cell(row=row, column=5, value=district)
            sheet.cell(row=row, column=6, value=address)
            sheet.cell(row=row, column=7, value=licence)
            sheet.cell(row=row, column=8, value=supervisor)
            print(row)
            row += 1
            i += 1        
    sh += 1
    wb.save('test.xlsx')


driver.quit()
